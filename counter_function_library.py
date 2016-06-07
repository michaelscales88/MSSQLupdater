import openpyxl
import datetime
from client_container import ClientContainer
from CONSTANTS import M_DRIVE_PATH


def get_sec(time_string):
    # returns time provided as as string in hours, minutes, sec
    try:
        h, m, s = [int(float(i)) for i in time_string.split(':')]
    except TypeError:
        return 0
    return convert_sec(h, m, s)
    # Converts hours, minutes, sec to seconds


def convert_sec(h, m, s):
    # Converts hours, minutes, sec to seconds
    return (3600 * int(h)) + (60 * int(m)) + int(s)


def convert_time_stamp(convert_seconds):
    # Converts seconds into a time stamp to be printed in a workbook.
    minutes, seconds = divmod(convert_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return "%d:%02d:%02d" % (hours, minutes, seconds)


def valid_input(column_position, row, ws, input_type):
    # Utility function for validating cells from each ws
    if input_type == 'N':
        try:
            return_value = int(ws['%s%d' % (column_position, row)].value)
        except TypeError:
            return_value = 0
        except ValueError:
            try:
                return_value = get_sec(ws['%s%d' % (column_position, row)].value)
            except ValueError:
                return_value = 0
    elif input_type == 'B':
        return_value = ws['%s%d' % (column_position, row)].value
    elif input_type == 'S':
        return_value = str(ws['%s%d' % (column_position, row)].value)
    elif input_type == 'D':
        try:
            return_value = ws['%s%d' % (column_position, row)].value
            return_value = datetime.datetime.strptime(return_value, '%m/%d/%Y')
        except TypeError:
            return_value = None
    else:
        raise ValueError("Invalid input type in valid_input")
    return return_value


def get_report_dates(input_string):
    while True:
        try:
            raw_date = input(input_string)
            final_date = datetime.datetime.strptime(raw_date, '%m/%d/%Y')
        except ValueError:
            print("Please enter the date in the correct format: MM/DD/YYYY")
            pass
        else:
            return final_date


def convert_day_month_year(date_to_split):
    day = int(date_to_split.day)
    if day < 10:
        day = '0%d' % day
    else:
        day = str(day)
    month = int(date_to_split.month)
    if month < 10:
        month = '0%d' % month
    else:
        month = str(month)
    year = str(date_to_split.year)
    return month, day, year


def check_value_in_range(min_range, max_range, work_sheet, max_or_min):
    return_row = 999
    if max_or_min == "Max":
        step_size = -1
    else:
        step_size = 1
    for row_number in range(min_range, max_range, step_size):
        working_cell = valid_input('A', row_number, work_sheet, 'S')
        working_cell = working_cell.split(' ')
        if working_cell[0].isdigit():
            return_row = row_number
            break
        else:
            pass
    return return_row


def find_min_max_rows(work_sheet):
    raw_max_row = work_sheet.max_row + 1
    min_row = check_value_in_range(1, raw_max_row, work_sheet, "Min")
    max_row = check_value_in_range(raw_max_row, 1, work_sheet, "Max")
    return min_row, max_row


def make_row_dictionary(work_sheet):
    return_dict = {}
    min_row, max_row = find_min_max_rows(work_sheet)
    for row_number in range(min_row, max_row + 1):
        cell_string = valid_input('A', row_number, work_sheet, 'S')
        try:
            split_cell = cell_string.split(' ')
        except ValueError:
            pass
        else:
            if split_cell[0].isdigit() and work_sheet.title == "REPORT":
                return_dict[split_cell[0]] = row_number + 1
            elif split_cell[0].isdigit() and work_sheet.title == "data":
                return_dict[split_cell[0]] = row_number + 2
            else:
                pass
    return return_dict


def open_client_report(start_date):
    try:
        client_report = []
        month, day, year = convert_day_month_year(start_date)
        output_file_to_inspect = openpyxl.load_workbook(M_DRIVE_PATH + '\\%s\\%s%s%s_Incoming DID Summary.xlsx' %
                                                        (year, month, day, year))
        data_page = output_file_to_inspect.get_sheet_by_name("data")
        report_page = output_file_to_inspect.get_sheet_by_name("REPORT")
        report_date = valid_input('AC', 1, data_page, 'D')
    except FileNotFoundError:
        client_report = None
        report_date = 0
        data_page = 0
        report_page = 0
        report_page_rows = {}
        data_page_rows = {}
    else:
        if report_date is None:
            report_date = start_date
        report_page_rows = make_row_dictionary(report_page)
        data_page_rows = make_row_dictionary(data_page)
        client_template = list(report_page_rows.keys())
        client_template = sorted(client_template)
        for client_name in client_template:
            client_container = ClientContainer(client_name)
            client_report.append(client_container)
    return (client_report,
            report_page_rows,
            data_page_rows,
            report_page,
            data_page,
            report_date)


def run_client_report(client_report, data_page_rows, report_page_rows,
                      data_page, report_page, report_date):
    for client in client_report:
        data_row_number = data_page_rows[client.get_name()]
        report_page_row = report_page_rows[client.get_name()]
        total_calls_presented = valid_input('C', data_row_number, data_page, 'N')
        total_calls_lost = valid_input('G', data_row_number, data_page, 'N')
        voice_mails_received = valid_input('F', report_page_row, report_page, 'N')
        average_incoming_duration = valid_input('L', data_row_number, data_page, 'N')
        average_wait_answered = valid_input('N', data_row_number, data_page, 'N')
        average_wait_lost = valid_input('P', data_row_number, data_page, 'N')
        ticker_less_than_15_sec = valid_input('R', data_row_number, data_page, 'N')
        ticker_less_than_30_sec = valid_input('S', data_row_number, data_page, 'N')
        ticker_less_than_45_sec = valid_input('T', data_row_number, data_page, 'N')
        ticker_less_than_60_sec = valid_input('U', data_row_number, data_page, 'N')
        ticker_greater_than_60_sec = valid_input('V', data_row_number, data_page, 'N')
        client.set_stats_for_client(report_date,
                                    total_calls_presented,
                                    total_calls_lost,
                                    voice_mails_received,
                                    average_incoming_duration,
                                    average_wait_answered,
                                    average_wait_lost,
                                    ticker_less_than_15_sec,
                                    ticker_less_than_30_sec,
                                    ticker_less_than_45_sec,
                                    ticker_less_than_60_sec,
                                    ticker_greater_than_60_sec)
    return client_report
