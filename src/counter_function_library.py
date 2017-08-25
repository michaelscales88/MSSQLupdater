import openpyxl as op
import datetime
from .client_container import ClientContainer
from .CONSTANTS import M_DRIVE_PATH, SELF_PATH


def get_sec(time_string):
    # returns time provided as as string in hours, minutes, sec
    try:
        h, m, s = [int(float(i)) for i in time_string.split(':')]
    except TypeError:
        return 0
    return convert_sec(h, m, s)
    # Converts hours, minutes, sec to seconds


def convert_sec(h, m, s):
    # Converts hours, minutes, sec to seconds
    return (3600 * int(h)) + (60 * int(m)) + int(s)


def convert_time_stamp(convert_seconds):
    # Converts seconds into a time stamp to be printed in a workbook.
    minutes, seconds = divmod(convert_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return "%d:%02d:%02d" % (hours, minutes, seconds)


def valid_input(column_position, row, ws, input_type):
    # Utility function for validating cells from each ws
    if input_type == 'N':
        try:
            return_value = int(ws['%s%d' % (column_position, row)].value)
        except TypeError:
            return_value = 0
        except ValueError:
            try:
                return_value = get_sec(ws['%s%d' % (column_position, row)].value)
            except ValueError:
                return_value = 0
    elif input_type == 'B':
        return_value = ws['%s%d' % (column_position, row)].value
    elif input_type == 'S':
        return_value = str(ws['%s%d' % (column_position, row)].value)
    elif input_type == 'D':
        try:
            return_value = ws['%s%d' % (column_position, row)].value
            return_value = datetime.datetime.strptime(return_value, '%m/%d/%Y')
        except TypeError:
            return_value = None
    else:
        raise ValueError("Invalid input type in valid_input")
    return return_value


def get_report_dates(input_string):
    while True:
        try:
            raw_date = input(input_string)
            final_date = datetime.datetime.strptime(raw_date, '%m/%d/%Y')
        except ValueError:
            print("Please enter the date in the correct format: MM/DD/YYYY")
            pass
        else:
            return final_date


def convert_day_month_year(date_to_split):
    day = int(date_to_split.day)
    if day < 10:
        day = '0%d' % day
    else:
        day = str(day)
    month = int(date_to_split.month)
    if month < 10:
        month = '0%d' % month
    else:
        month = str(month)
    year = str(date_to_split.year)
    return month, day, year


def check_value_in_range(min_range, max_range, work_sheet, max_or_min):
    return_row = 999
    if max_or_min == "Max":
        step_size = -1
    else:
        step_size = 1
    for row_number in range(min_range, max_range, step_size):
        working_cell = valid_input('A', row_number, work_sheet, 'S')
        working_cell = working_cell.split(' ')
        if working_cell[0].isdigit():
            return_row = row_number
            break
        else:
            pass
    return return_row


def find_min_max_rows(work_sheet):
    raw_max_row = work_sheet.max_row + 1
    min_row = check_value_in_range(1, raw_max_row, work_sheet, "Min")
    max_row = check_value_in_range(raw_max_row, 1, work_sheet, "Max")
    return min_row, max_row


def make_row_dictionary(work_sheet):
    return_dict = {}
    min_row, max_row = find_min_max_rows(work_sheet)
    for row_number in range(min_row, max_row + 1):
        cell_string = valid_input('A', row_number, work_sheet, 'S')
        try:
            split_cell = cell_string.split(' ')
        except ValueError:
            pass
        else:
            if split_cell[0].isdigit() and work_sheet.title == "REPORT":
                return_dict[split_cell[0]] = row_number + 1
            elif split_cell[0].isdigit() and work_sheet.title == "data":
                return_dict[split_cell[0]] = row_number + 2
            else:
                pass
    return return_dict


def open_client_report(start_date):
    try:
        client_report = []
        month, day, year = convert_day_month_year(start_date)
        output_file_to_inspect = op.load_workbook(M_DRIVE_PATH + '\\%s\\%s%s%s_Incoming DID Summary.xlsx' %
                                                  (year, month, day, year))
        data_page = output_file_to_inspect.get_sheet_by_name("data")
        report_page = output_file_to_inspect.get_sheet_by_name("REPORT")
        report_date = valid_input('AC', 1, data_page, 'D')
    except FileNotFoundError:
        client_report = None
        report_date = 0
        data_page = 0
        report_page = 0
        report_page_rows = {}
        data_page_rows = {}
    else:
        if report_date is None:
            report_date = start_date
        report_page_rows = make_row_dictionary(report_page)
        data_page_rows = make_row_dictionary(data_page)
        client_template = sorted(report_page_rows.keys())
        for client_name in client_template:
            client_container = ClientContainer(client_name)
            client_report.append(client_container)
    return (client_report,
            report_page_rows,
            data_page_rows,
            report_page,
            data_page,
            report_date)


def run_client_report(client_report, data_page_rows, report_page_rows,
                      data_page, report_page, report_date):
    for client in client_report:
        data_row_number = data_page_rows[client.get_name()]
        report_page_row = report_page_rows[client.get_name()]
        total_calls_presented = valid_input('C', data_row_number, data_page, 'N')
        total_calls_lost = valid_input('G', data_row_number, data_page, 'N')
        voice_mails_received = valid_input('F', report_page_row, report_page, 'N')
        average_incoming_duration = valid_input('L', data_row_number, data_page, 'N')
        average_wait_answered = valid_input('N', data_row_number, data_page, 'N')
        average_wait_lost = valid_input('P', data_row_number, data_page, 'N')
        ticker_less_than_15_sec = valid_input('R', data_row_number, data_page, 'N')
        ticker_less_than_30_sec = valid_input('S', data_row_number, data_page, 'N')
        ticker_less_than_45_sec = valid_input('T', data_row_number, data_page, 'N')
        ticker_less_than_60_sec = valid_input('U', data_row_number, data_page, 'N')
        ticker_greater_than_60_sec = valid_input('V', data_row_number, data_page, 'N')
        client.set_stats_for_client(report_date,
                                    total_calls_presented,
                                    total_calls_lost,
                                    voice_mails_received,
                                    average_incoming_duration,
                                    average_wait_answered,
                                    average_wait_lost,
                                    ticker_less_than_15_sec,
                                    ticker_less_than_30_sec,
                                    ticker_less_than_45_sec,
                                    ticker_less_than_60_sec,
                                    ticker_greater_than_60_sec)
    return client_report


def make_head_node():
    head_node = {}
    client_book = op.load_workbook(SELF_PATH + '\\bin\\client_list_file.xlsx')
    client_page = client_book.get_sheet_by_name('Sheet1')
    for row in range(1, client_page.max_row + 1):
        client_name = str(client_page['A%d' % row].value)
        client = ClientContainer(client_name)
        head_node[client_name] = client
    return head_node


def get_client_dictionary():
    wb = op.load_workbook(SELF_PATH + '\\bin\\client_list_file.xlsx')
    ws = wb.get_sheet_by_name("Sheet1")
    return_dict = {}
    for row_number in range(1, ws.max_row + 1):
        client_name = valid_input('B', row_number, ws, 'S')
        client_number = valid_input('A', row_number, ws, 'N')
        return_dict[client_number] = client_name
    return return_dict


def create_output_report_xlsx(client_report):
    today_date = datetime.datetime.today()
    new_xlsx_output_file = op.Workbook()
    client_dict = get_client_dictionary()
    report_page = new_xlsx_output_file.create_sheet("REPORT", 0)
    report_page['A1'] = "{0}".format(today_date.strftime("%A %m/%d/%Y"))
    report_page['B1'] = "I/C Presented"
    report_page['C1'] = "I/C Answered"
    report_page['D1'] = "I/C Lost"
    report_page['E1'] = "Voice Mails"
    report_page['F1'] = "Incoming Answered (%)"
    report_page['G1'] = "Incoming Lost (%)"
    report_page['H1'] = "Average Incoming Duration"
    report_page['I1'] = "Average Wait Answered"
    report_page['J1'] = "Average Wait Lost"
    report_page['K1'] = "Calls Ans Within 15"
    report_page['L1'] = "Calls Ans Within 30"
    report_page['M1'] = "Calls Ans Within 45"
    report_page['N1'] = "Calls Ans Within 60"
    report_page['O1'] = "Calls Ans Within 999"
    report_page['P1'] = "Call Ans + 999"
    report_page['Q1'] = "Longest Waiting Answered"
    report_page['R1'] = "PCA"

    total_calls_received = 0
    total_calls_answered = 0
    total_calls_lost = 0
    total_voice_mails_received = 0
    average_incoming_duration = 0
    average_wait_answered = 0
    average_wait_lost = 0
    calls_within_15_seconds = 0
    calls_within_30_seconds = 0
    calls_within_45_seconds = 0
    calls_within_60_seconds = 0
    calls_greater_than_60_seconds = 0
    calls_answered_plus = 0
    longest_waiting_answered = 0
    divisor = 0

    client_names = sorted(client_report.keys())
    for row_number, client_name in enumerate(client_names):
        client = client_report[client_name]
        try:
            client_name_verbose = client_dict[int(client.get_name())]
        except KeyError:
            client_name_verbose = ''
        adjusted_row = row_number + 2
        report_page['A%d' % adjusted_row] = "{0} {1}".format(client_name, client_name_verbose)
        report_page['B%d' % adjusted_row] = client.get_total_calls()
        if client.get_total_calls() != 0:
            divisor += 1
        total_calls_received += client.get_total_calls()
        report_page['C%d' % adjusted_row] = client.get_calls_answered_true()
        total_calls_answered += client.get_calls_answered_true()
        report_page['D%d' % adjusted_row] = client.get_true_calls_lost()
        total_calls_lost += client.get_true_calls_lost()
        report_page['E%d' % adjusted_row] = client.get_voice_mails()
        total_voice_mails_received += client.get_voice_mails()
        report_page['F%d' % adjusted_row] = client.get_calls_answered_percentage()
        report_page['G%d' % adjusted_row] = client.get_calls_lost_percentage()
        report_page['H%d' % adjusted_row] = convert_time_stamp(client.get_average_duration())
        average_incoming_duration += client.get_average_duration()
        report_page['I%d' % adjusted_row] = convert_time_stamp(client.get_average_wait_answered())
        average_wait_answered += client.get_average_wait_answered()
        report_page['J%d' % adjusted_row] = convert_time_stamp(client.get_average_wait_lost())
        average_wait_lost += client.get_average_wait_lost()
        (call_15sec,
         call_30sec,
         call_45sec,
         call_60sec,
         call_g60sec) = client.get_ticker()
        report_page['K%d' % adjusted_row] = call_15sec
        calls_within_15_seconds += call_15sec
        report_page['L%d' % adjusted_row] = call_30sec
        calls_within_30_seconds += call_30sec
        report_page['M%d' % adjusted_row] = call_45sec
        calls_within_45_seconds += call_45sec
        report_page['N%d' % adjusted_row] = call_60sec
        calls_within_60_seconds += call_60sec
        report_page['O%d' % adjusted_row] = call_g60sec
        calls_greater_than_60_seconds += call_g60sec
        report_page['P%d' % adjusted_row] = 0
        calls_answered_plus += 0
        # report_page['Q%d' % adjusted_row] = convert_time_stamp(client.get_longest_answered())
        # if client.get_longest_answered() > longest_waiting_answered:
        #     longest_waiting_answered = client.get_longest_answered()
        report_page['R%d' % adjusted_row] = client.get_pca_percentage()

    last_row = report_page.max_row + 1
    report_page['A%d' % last_row] = "Total"
    report_page['B%d' % last_row] = total_calls_received
    report_page['C%d' % last_row] = total_calls_answered
    report_page['D%d' % last_row] = total_calls_lost
    report_page['E%d' % last_row] = total_voice_mails_received
    report_page['F%d' % last_row] = "{0}%".format(round((total_calls_answered / total_calls_received) * 100, 2))
    combined_lost_calls = total_calls_lost + total_voice_mails_received
    report_page['G%d' % last_row] = "{0}%".format(round((combined_lost_calls / total_calls_received) * 100, 2))
    report_page['H%d' % last_row] = convert_time_stamp(average_incoming_duration / divisor)
    report_page['I%d' % last_row] = convert_time_stamp(average_wait_answered / divisor)
    report_page['J%d' % last_row] = convert_time_stamp(average_wait_lost / divisor)
    report_page['K%d' % last_row] = calls_within_15_seconds
    report_page['L%d' % last_row] = calls_within_30_seconds
    report_page['M%d' % last_row] = calls_within_45_seconds
    report_page['N%d' % last_row] = calls_within_60_seconds
    report_page['O%d' % last_row] = calls_greater_than_60_seconds
    report_page['P%d' % last_row] = calls_answered_plus
    report_page['Q%d' % last_row] = convert_time_stamp(longest_waiting_answered)
    total_pca = ((calls_within_15_seconds + calls_within_30_seconds) / total_calls_answered) * 100
    report_page['R%d' % last_row] = "{0}%".format(round(total_pca, 2))

    new_xlsx_output_file.save('%s\\output\\temp_file.xlsx' % SELF_PATH)
